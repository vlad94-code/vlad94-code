"""Ежедневная выгрузка остатков и приходов в Excel.

Заменяет файл, который специалист по документообороту каждый рабочий день
выгружает из 1С и рассылает сотрудникам — настолько, насколько это позволяет
CNC API. Из четырёх синхронизируемых источников собираются артикул,
наименование, вид номенклатуры, единица, цена, складской остаток и приходы,
разложенные по датам поступления — так же, как в исходном файле 1С.

Важно про число в колонке «Доступно»: остаток из API уже очищен от резерва.
Сверка с дневным файлом 1С от 24.08: из 249 позиций с резервом у 243 значение
API совпало с колонкой «Доступно» и лишь у одной — с «В наличии» (A034034:
249 в наличии, 223 в резерве, 26 доступно — API даёт ровно 26). Поэтому
колонка названа так же, как в 1С, и продавать по ней можно напрямую.

Чего в выгрузке нет — самого резерва. Он нужен не для арифметики, а для
разговора: видя «доступно 0, в резерве 10», менеджер идёт спрашивать,
не протух ли чужой резерв — а с горячим заказом это решает сделку. По той
же причине нет колонок «Отгружается», «К обеспечению», «Дефицит» и
«Излишек»: это данные управления заказами, которых в наших эндпоинтах нет.
"""
from __future__ import annotations

import datetime as dt
from typing import Any, Iterable, Mapping

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

TITLE = "Остатки и приходы CNC Electric"
RESERVE_WARNING = (
    "«Доступно» — то, что можно продавать прямо сейчас. Сам резерв здесь не "
    "показан: если доступно 0, товар может лежать в чужом резерве — посмотрите в 1С, "
    "может, он уже неактуален."
)
SOURCE_NOTE = "Источник: CNC API (products, prices, stock-balances, GoodsInTransit)."
FIXED_COLUMNS = ["Артикул", "Вид номенклатуры", "Наименование", "Ед. изм.", "Цена", "Доступно"]
COLUMN_WIDTHS = [12, 26, 52, 9, 12, 11]


def _number(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(str(value).replace(" ", "").replace("\xa0", "").replace(",", "."))
    except ValueError:
        return None


def _tidy(value: float | None) -> float | int | None:
    """Пустая клетка вместо нуля — как в исходном файле 1С, где отсутствие
    остатка не пишут нулём, чтобы глаз цеплялся только за реальные числа."""
    if not value:
        return None
    return int(value) if float(value).is_integer() else round(value, 2)


def _format_date(raw: str) -> str:
    try:
        return dt.date.fromisoformat(raw).strftime("%d.%m.%Y")
    except (TypeError, ValueError):
        return str(raw or "")


def report_filename(when: dt.datetime) -> str:
    return "Остатки CNC %s.xlsx" % when.strftime("%d.%m.%Y")


def build_report(
    catalog: Iterable[Mapping[str, Any]],
    prices: Mapping[str, Any],
    stock: Mapping[str, float],
    transit: Mapping[str, list[tuple[float, str]]],
    when: dt.datetime,
) -> Workbook:
    """Собирает книгу. Строка появляется только у товаров, где есть остаток или
    приход: остальные 10 000 позиций каталога в складской сводке — шум."""
    by_code = {
        str(row.get("vendor_code", "")).upper(): row
        for row in catalog
        if isinstance(row, Mapping) and row.get("vendor_code")
    }
    dates = sorted({date for arrivals in transit.values() for _, date in arrivals if date})
    articles = sorted({code for code in stock if stock.get(code)} | {code for code in transit if transit.get(code)})

    book = Workbook()
    sheet = book.active
    sheet.title = "Остатки"

    sheet.append([TITLE])
    sheet["A1"].font = Font(bold=True, size=13)
    sheet.append(["Сформировано: %s" % when.strftime("%d.%m.%Y %H:%M")])
    sheet.append([SOURCE_NOTE])
    sheet.append([RESERVE_WARNING])
    sheet["A4"].font = Font(bold=True)
    sheet["A4"].alignment = Alignment(wrap_text=False)
    sheet.append([])

    header = FIXED_COLUMNS + ["Поступит %s" % _format_date(date) for date in dates] + ["Всего поступит"]
    sheet.append(header)
    for cell in sheet[sheet.max_row]:
        cell.font = Font(bold=True)
    header_row = sheet.max_row

    for code in articles:
        product = by_code.get(code, {})
        arrivals = {date: 0.0 for date in dates}
        for quantity, date in transit.get(code, []):
            if date in arrivals:
                arrivals[date] += quantity or 0.0
        total = sum(arrivals.values())
        sheet.append(
            [
                code,
                str(product.get("type_item", "") or ""),
                str(product.get("name", "") or ""),
                str(product.get("unit_name", "") or "шт"),
                _tidy(_number(prices.get(code))),
                _tidy(stock.get(code)),
            ]
            + [_tidy(arrivals[date]) for date in dates]
            + [_tidy(total)]
        )

    for index, width in enumerate(COLUMN_WIDTHS, start=1):
        sheet.column_dimensions[sheet.cell(row=header_row, column=index).column_letter].width = width
    for index in range(len(FIXED_COLUMNS) + 1, len(header) + 1):
        sheet.column_dimensions[sheet.cell(row=header_row, column=index).column_letter].width = 17
    sheet.freeze_panes = sheet.cell(row=header_row + 1, column=1)
    return book
