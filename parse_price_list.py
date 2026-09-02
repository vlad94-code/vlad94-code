# -*- coding: utf-8 -*-
"""
parse_price_list.py

Парсер официального прайс-листа CNC Electric (xlsx) в единую нормализованную
товарную матрицу (JSON) для бота.

Логика опирается на реальную структуру файла:
  - каждый лист = категория оборудования (модульное, силовое, коммутация, DC,
    реле, светосигнальная аппаратура, блоки питания, измерительное оборудование)
  - строки внутри листа делятся на:
      * товары (обычные позиции)
      * принадлежности/аксессуары — определяются по заполненной колонке
        "Аксессуары"/"Подтип"=Принадлежность; поле "Серия" в этом случае
        содержит список серий через "/", к которым подходит аксессуар
  - товар, снятый с производства, но ещё есть на складе — выделен заливкой
    ячейки в колонке "Артикул" цветом FF0070C0 (это подтверждено и описано
    в примечаниях листа "Главная")
  - отдельный лист "Выведено из ассортимента" содержит товары, полностью
    снятые с производства
  - код артикула однозначно кодирует категорию первыми буквами (A/B/C/EN/D/E/F/G)

Результат — products_matrix.json, список плоских объектов с общими полями
(article, category_code, category_name, type, name, price, series, ...)
и словарём specs с остальными характеристиками листа (они разные для каждой
категории оборудования).
"""

import json
import re
import sys
from datetime import date
from openpyxl import load_workbook

SRC = "/mnt/user-data/uploads/8865_284_Prays-list-CNC_V3_8_1-28072026-_obshchiy_.xlsx"
OUT = "/mnt/user-data/outputs/products_matrix.json"

DISCONTINUED_FILL = "FF0070C0"  # синяя заливка = снято с производства, есть на складе

# Листы -> код категории (совпадает с префиксом артикула) и статус
CATEGORY_SHEETS = {
    "A (модульное)": ("A", "active"),
    "B(силовое)": ("B", "active"),
    "С(коммутация)": ("C", "active"),
    "EN (DC)": ("EN", "active"),
    "D (реле)": ("D", "active"),
    "E (светосигн)": ("E", "active"),
    "F(блоки питания)": ("F", "active"),
    "G (измерительное оборудование)": ("G", "active"),
    "Выведено из ассортимента": (None, "discontinued"),
}

# Поля, которые вытаскиваем в отдельные "верхнеуровневые" атрибуты товара.
# Всё, что не попало в этот список, уходит в specs{}.
COMMON_FIELD_MAP = {
    "Артикул": "article",
    "Тип": "type",
    "Описание": "name",
    "Аксессуары": "accessory_kind",   # если заполнено -> это аксессуар
    "Подтип": "subtype",
    "Серия": "series",
    "Типоразмер": "size",
    "Категория оборудования": "category_name",
    "Габариты коробки(см) ": "box_dims_cm",
    "Габариты коробки(см)": "box_dims_cm",
    "Вес брутто (кг)": "weight_kg_gross",
    "Кратность упаковки (шт)": "pack_multiplicity",
}

PRICE_HEADER_RE = re.compile(r"Цена\s+с\s+(\d{2})\.(\d{2})\.(\d{4})")


def find_price_field(headers):
    for h in headers:
        if h and h.startswith("Цена"):
            m = PRICE_HEADER_RE.search(h)
            price_date = None
            if m:
                d, mo, y = m.groups()
                price_date = f"{y}-{mo}-{d}"
            return h, price_date
    return None, None


def clean(v):
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip()
        return v if v else None
    return v


def parse_sheet(ws, sheet_name, cat_code, status_default):
    headers = [c.value for c in ws[1]]
    price_field, price_date = find_price_field(headers)

    records = []
    for row_idx, row in enumerate(
        ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False), start=2
    ):
        values = [c.value for c in row]
        article = clean(values[0])
        if not article:
            continue

        rec = {
            "article": str(article),
            "category_code": cat_code or re.match(r"^[A-Za-zА-Яа-я]+", str(article)).group(),
            "sheet": sheet_name,
            "price": None,
            "price_date": price_date,
            "series": None,
            "size": None,
            "type": None,
            "name": None,
            "category_name": None,
            "accessory_kind": None,
            "is_accessory": False,
            "compatible_series": [],
            "status": status_default,
            "specs": {},
        }

        specs = {}
        for h, v in zip(headers, values):
            if h is None:
                continue
            v = clean(v)
            if h == price_field:
                rec["price"] = float(v) if isinstance(v, (int, float)) else v
                continue
            mapped = COMMON_FIELD_MAP.get(h)
            if mapped:
                rec[mapped] = v
            else:
                if v is not None:
                    specs[h] = v

        rec["specs"] = specs

        if rec["accessory_kind"]:
            rec["is_accessory"] = True
            if rec["series"]:
                rec["compatible_series"] = [s.strip() for s in str(rec["series"]).split("/") if s.strip()]

        # discontinued-but-in-stock detection via cell fill on article column
        if status_default == "active":
            cell = ws.cell(row=row_idx, column=1)
            fill = cell.fill.fgColor.rgb if cell.fill and cell.fill.fgColor else None
            if fill == DISCONTINUED_FILL:
                rec["status"] = "discontinued_in_stock"

        records.append(rec)

    return records


def main():
    wb = load_workbook(SRC, data_only=True)
    all_records = []
    stats = {}

    for sheet_name, (cat_code, status_default) in CATEGORY_SHEETS.items():
        if sheet_name not in wb.sheetnames:
            print(f"WARNING: sheet not found: {sheet_name}", file=sys.stderr)
            continue
        ws = wb[sheet_name]
        recs = parse_sheet(ws, sheet_name, cat_code, status_default)
        all_records.extend(recs)
        stats[sheet_name] = len(recs)

    # sanity: unique article check
    seen = {}
    dupes = []
    for r in all_records:
        key = (r["article"], r["status"] == "discontinued")
        if key in seen:
            dupes.append(r["article"])
        seen[key] = True

    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(all_records, f, ensure_ascii=False, indent=2)

    print("Records per sheet:")
    for k, v in stats.items():
        print(f"  {k}: {v}")
    print("TOTAL:", len(all_records))
    print("Possible duplicate articles (within same status group):", len(dupes))
    print("Written to:", OUT)


if __name__ == "__main__":
    main()