"""passport_links — разбор таблицы «Серия / Ссылка на паспорт», импорт в
изолированную БД и подбор ссылки под конкретный товар.

Таблица ведётся вручную и приходит наполовину пустой: часть строк — нули,
часть — неразвёрнутые формулы =VLOOKUP(...). Разбор обязан молча оставлять
только настоящие ссылки, иначе бот пришлёт менеджеру «0» вместо паспорта.
Боевой data/knowledge.db тесты не трогают."""
import sqlite3

import openpyxl
import pytest

import passport_links
from passport_links import PassportLink

_URL = "https://cncrussia.com/uploads/passports/ycw3_pasport.pdf"


def _workbook(tmp_path, rows) -> "object":
    path = tmp_path / "passports.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["Серия", "Ссылки на паспорта", "Статус"])
    for row in rows:
        sheet.append(list(row))
    workbook.save(path)
    workbook.close()
    return path


@pytest.fixture
def isolated_db(tmp_path, monkeypatch):
    db_path = tmp_path / "passports_test.db"
    monkeypatch.setattr(passport_links, "DB_PATH", db_path)

    def _make_document(document_id: int, version: int = 1) -> None:
        connection = sqlite3.connect(db_path)
        try:
            connection.execute(
                "CREATE TABLE IF NOT EXISTS documents (id INTEGER PRIMARY KEY, version INTEGER, uploaded_at TEXT)"
            )
            connection.execute(
                "INSERT INTO documents (id, version, uploaded_at) VALUES (?, ?, ?)",
                (document_id, version, "2026-08-27T10:00:00"),
            )
            connection.commit()
        finally:
            connection.close()

    return _make_document


# --- Разбор файла ------------------------------------------------------------

def test_parse_keeps_every_series_and_marks_the_missing_links(tmp_path):
    """Строки без ссылки («0», формула, пусто) не выбрасываются: серия, у
    которой паспорта ещё нет, обязана остаться в базе — иначе она получит
    паспорт более общей серии из наименования (см. блокировку ниже)."""
    path = _workbook(tmp_path, [
        ("YCB9-125", "https://cncrussia.com/uploads/passports/ycb9-125_pasport.pdf", "ок"),
        ("YCBZ-40", "0", "сделать"),
        ("YCMV", "=VLOOKUP(A23,'[1]A-Модульное оборуд.'!$B$2:$I$48,4,0)", "сделать"),
        ("YCW1", None, "сделать"),
    ])
    links = passport_links.parse_workbook(path)
    assert [link.series for link in links] == ["YCB9-125", "YCBZ-40", "YCMV", "YCW1"]
    assert links[0].url.endswith("ycb9-125_pasport.pdf")
    assert links[0].status == "ок"
    assert [link.url for link in links[1:]] == [None, None, None]


def test_parse_normalises_series_to_upper_case(tmp_path):
    path = _workbook(tmp_path, [
        (" ycb9-125 ", "https://cncrussia.com/uploads/passports/ycb9-125_pasport.pdf", "ок"),
    ])
    assert passport_links.parse_workbook(path)[0].series == "YCB9-125"


def test_parse_finds_columns_by_header_not_by_position(tmp_path):
    """Столбцы в присылаемом файле могут переехать — шапка остаётся."""
    path = tmp_path / "moved.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["Статус", "Серия", "Ссылки на паспорта"])
    sheet.append(["ок", "YCW3", "https://cncrussia.com/uploads/passports/ycw3_pasport.pdf"])
    workbook.save(path)
    workbook.close()
    links = passport_links.parse_workbook(path)
    assert links[0].series == "YCW3"
    assert links[0].url.endswith("ycw3_pasport.pdf")


def test_parse_refuses_a_workbook_without_the_expected_headers(tmp_path):
    path = tmp_path / "wrong.xlsx"
    workbook = openpyxl.Workbook()
    workbook.active.append(["Артикул", "Цена"])
    workbook.save(path)
    workbook.close()
    with pytest.raises(ValueError):
        passport_links.parse_workbook(path)


# --- Хранение ----------------------------------------------------------------

def test_import_stores_links(isolated_db):
    isolated_db(document_id=1)
    inserted = passport_links.import_links(1, [
        PassportLink("YCW3", "https://cncrussia.com/uploads/passports/ycw3_pasport.pdf", "ок"),
        PassportLink("YCM3", "https://cncrussia.com/uploads/passports/ycm3_pasport.pdf", "ок"),
    ])
    assert inserted == 2
    assert passport_links.link_for("YCW3").url.endswith("ycw3_pasport.pdf")


def test_reimport_drops_the_previous_version(isolated_db):
    isolated_db(document_id=1)
    isolated_db(document_id=2, version=2)
    passport_links.import_links(1, [
        PassportLink("YCW3", "https://cncrussia.com/uploads/passports/old.pdf", "ок"),
        PassportLink("YCM3", "https://cncrussia.com/uploads/passports/ycm3_pasport.pdf", "ок"),
    ])
    passport_links.import_links(2, [
        PassportLink("YCW3", "https://cncrussia.com/uploads/passports/new.pdf", "ок"),
    ])
    assert passport_links.link_for("YCW3").url.endswith("new.pdf")
    assert passport_links.link_for("YCM3") is None


def test_series_listed_twice_does_not_break_the_import(isolated_db):
    isolated_db(document_id=1)
    inserted = passport_links.import_links(1, [
        PassportLink("YCHGL", "https://cncrussia.com/uploads/passports/ychglychglz1_pasport.pdf", "замена"),
        PassportLink("YCHGL", "https://cncrussia.com/uploads/passports/ychglychglz1_pasport.pdf", "замена"),
    ])
    assert inserted == 1


# --- Подбор ссылки под товар -------------------------------------------------

@pytest.fixture
def loaded(isolated_db):
    isolated_db(document_id=1)
    passport_links.import_links(1, [
        PassportLink("YCB9-125", "https://cncrussia.com/uploads/passports/ycb9-125_pasport.pdf", "ок"),
        PassportLink("YCB9ZF-100AP", "https://cncrussia.com/uploads/passports/ycb9zf-100ap_pasport_a.pdf", "ок"),
        PassportLink("AFDD", "https://cncrussia.com/uploads/passports/pasport_AFDD.pdf", "ок"),
        PassportLink("YC7VA", "https://cncrussia.com/uploads/passports/YC7VA_Instructions(1).pdf", "ок"),
        PassportLink("YC7VAN", "https://cncrussia.com/uploads/passports/YC7VAN_Instructions.250610.pdf", "ок"),
    ])


def test_exact_series_wins(loaded):
    assert passport_links.link_for("YCB9-125").url.endswith("ycb9-125_pasport.pdf")


def test_series_is_matched_case_insensitively(loaded):
    assert passport_links.link_for(" ycb9-125 ") is not None


def test_unknown_series_has_no_link(loaded):
    assert passport_links.link_for("YCB8-63") is None
    assert passport_links.link_for("") is None


def test_series_of_the_catalogue_is_shorter_than_in_the_table(loaded):
    """В каталоге серия «YCB9ZF», в таблице — «YCB9ZF-100AP»: точного
    совпадения нет, но наименование товара называет исполнение полностью."""
    link = passport_links.link_for(
        "YCB9ZF", "Автоматический выключатель YCB9ZF-100AP-2P 4G + 485, 32A"
    )
    assert link.url.endswith("ycb9zf-100ap_pasport_a.pdf")


def test_series_of_the_catalogue_is_longer_than_in_the_table(loaded):
    """В каталоге «AFDD L1», в таблице — «AFDD»."""
    link = passport_links.link_for("AFDD L1", "Устройство защиты от дугового пробоя AFDD L1 C6А 2P 6kA 36мА")
    assert link.url.endswith("pasport_AFDD.pdf")


def test_two_products_of_one_catalogue_series_get_their_own_passports(loaded):
    """У YC7VA-3 и YC7VAN-1 серия каталога одна — «YC7»; различить их можно
    только по наименованию, и перепутать паспорта здесь нельзя."""
    vа = passport_links.link_for("YC7", "Реле контроля напряжения YC7VA-3 63A AC380")
    van = passport_links.link_for("YC7", "Реле контроля напряжения YC7VAN-1 40A AC220")
    assert vа.url.endswith("YC7VA_Instructions(1).pdf")
    assert van.url.endswith("YC7VAN_Instructions.250610.pdf")


def test_a_name_match_never_beats_an_exact_series(loaded):
    """Наименование почти всегда содержит и серию каталога, и что-то ещё —
    точное совпадение по серии должно оставаться главным."""
    link = passport_links.link_for("YCB9-125", "Автоматический выключатель YCB9-125 AFDD-совместимый")
    assert link.url.endswith("ycb9-125_pasport.pdf")


def test_a_series_token_inside_a_longer_word_is_not_a_match(loaded):
    assert passport_links.link_for("XXX", "Клеммник AFDDX-12 на din-рейку") is None


# --- Чужой паспорт прислать нельзя -------------------------------------------

@pytest.fixture
def loaded_with_gaps(isolated_db):
    """Так выглядит настоящая таблица: у YCM8 паспорт есть, у исполнения
    YCM8 PV — «сделать», у CJX2-K есть, а тепловому реле JR28 он не подходит,
    хотя контактор назван в наименовании реле как совместимый."""
    isolated_db(document_id=1)
    passport_links.import_links(1, [
        PassportLink("YCM8", "https://cncrussia.com/uploads/passports/ycm8_pasport.pdf", "ок"),
        PassportLink("YCM8 PV", None, "сделать"),
        PassportLink("CJX2-K", "https://cncrussia.com/uploads/passports/cjx2-k_pasport_b.pdf", "замена"),
        PassportLink("YC", "https://cncrussia.com/uploads/passports/yc-9sy-3e_h_lcd.pdf", "ок"),
    ])


def test_a_series_without_a_link_does_not_inherit_the_parent_passport(loaded_with_gaps):
    link = passport_links.link_for("YCM8 PV", "Выключатель в литом корпусе DC YCM8-250S PV 2P 250A DC1000")
    assert link is None


def test_a_series_named_in_the_catalogue_but_not_in_the_table_gets_nothing(loaded_with_gaps):
    """JR28 — тепловое реле; CJX2-K в наименовании это контактор, к которому
    реле подходит. Его паспорт для реле — чужой документ."""
    link = passport_links.link_for("JR28", "Тепловое реле JR28-11.5 1.6-2.5A CJX2-K")
    assert link is None


def test_a_product_without_a_series_gets_nothing(loaded_with_gaps):
    """Без серии каталога проверить родство не с чем, а имя может называть
    совместимый аппарат, а не сам товар."""
    assert passport_links.link_for("", "Контактор CJX2-K 09A 3P") is None


def test_count_counts_only_series_that_actually_have_a_passport(loaded_with_gaps):
    assert passport_links.count() == 3


# --- Кэш file_id -------------------------------------------------------------
# Паспорта у CNC весят 5–13 МБ, а сайт отдаёт их медленно (замер 27.08.2026:
# 770 КБ за 54 с). Второй раз тот же файл качать нельзя: Telegram умеет
# пересылать уже загруженный документ по file_id мгновенно.

def test_file_id_is_remembered_per_url(isolated_db):
    passport_links.remember_file_id(_URL, "BQACAgIAAxkBAAI")
    assert passport_links.cached_file_id(_URL) == "BQACAgIAAxkBAAI"


def test_unknown_url_has_no_file_id(isolated_db):
    assert passport_links.cached_file_id(_URL) is None


def test_a_file_id_survives_a_new_version_of_the_table(isolated_db):
    """Перезалили таблицу — ссылки те же, перекачивать их незачем."""
    isolated_db(document_id=1)
    isolated_db(document_id=2, version=2)
    passport_links.import_links(1, [PassportLink("YCW3", _URL, "ок")])
    passport_links.remember_file_id(_URL, "BQACAgIAAxkBAAI")
    passport_links.import_links(2, [PassportLink("YCW3", _URL, "ок")])
    assert passport_links.cached_file_id(_URL) == "BQACAgIAAxkBAAI"


def test_a_stale_file_id_can_be_forgotten(isolated_db):
    passport_links.remember_file_id(_URL, "BQACAgIAAxkBAAI")
    passport_links.forget_file_id(_URL)
    assert passport_links.cached_file_id(_URL) is None


def test_reuploading_the_same_url_replaces_the_file_id(isolated_db):
    passport_links.remember_file_id(_URL, "СТАРЫЙ")
    passport_links.remember_file_id(_URL, "НОВЫЙ")
    assert passport_links.cached_file_id(_URL) == "НОВЫЙ"
