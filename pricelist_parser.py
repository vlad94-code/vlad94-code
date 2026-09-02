"""Parse the official CNC Electric price-list workbook (xlsx) into a flat,
normalised list of ParsedItem — one entry per product/accessory row across
every equipment-category sheet.

This is a structural import (type/series/size/accessory-compatibility), not
a pricing one: the "Цена ..." column is read only to be skipped, never
stored anywhere (see ARCHITECTURE.md §5, критическая оговорка). Price and
stock stay exclusively the CNC API's job (catalog_search.py).

Ports the sheet knowledge from the earlier, unintegrated parse_price_list.py
draft (written against sandbox paths that don't exist in this repo, never
imported by anything here) — with one important correction: accessory rows
are detected via the "Тип" column, not "Аксессуары"/"Подтип". "Тип" is
present identically on every sheet and reliably contains "Принадлеж..."
for accessory rows even on sheets (e.g. "E (светосигн)") that have no
"Аксессуары" column at all — verified directly against the real workbook.
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

# Sheet name -> (category_code, default status for rows on that sheet).
# category_code is a fallback label only (also re-derived from the article's
# own prefix below) — CNC's sheet naming isn't a strict schema contract.
CATEGORY_SHEETS: dict[str, tuple[str | None, str]] = {
    "A (модульное)": ("A", "active"),
    "B(силовое)": ("B", "active"),
    "С(коммутация)": ("C", "active"),
    "EN (DC)": ("EN", "active"),
    "D (реле)": ("D", "active"),
    "E (светосигн)": ("E", "active"),
    "F(блоки питания)": ("F", "active"),
    "G (измерительное оборудование)": ("G", "active"),
    "Выведено из ассортимента": (None, "discontinued"),
}

# Columns promoted to their own ParsedItem field; everything else (minus the
# price column, which is dropped entirely) goes into .specs.
_COMMON_FIELD_MAP = {
    "Артикул": "article",
    "Тип": "type_field",
    "Описание": "name",
    "Серия": "series",
    "Типоразмер": "size",
}

_PRICE_HEADER_RE = re.compile(r"^Цена\b", re.I)
_ACCESSORY_TYPE_RE = re.compile(r"Принадлеж", re.I)

# Blue fill on the "Артикул" cell = discontinued but still in stock (per the
# workbook's own "Главная" sheet notes). Confirmed present in the real file
# (8865_248_Prays-list-CNC_V3.8.2.xlsx): 54 such cells on sheet B alone.
_DISCONTINUED_IN_STOCK_FILL = "FF0070C0"


@dataclass(frozen=True)
class ParsedItem:
    article: str
    sheet: str
    category_code: str | None
    type_field: str
    name: str | None
    series: str | None                 # single series; None for accessory rows
    size: str | None
    is_accessory: bool
    status: str                        # active | discontinued | discontinued_in_stock
    specs: dict[str, str] = field(default_factory=dict)
    compatible_series: list[str] = field(default_factory=list)


def _clean(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _category_code_for_article(article: str, sheet_default: str | None) -> str | None:
    if sheet_default:
        return sheet_default
    match = re.match(r"^[A-Za-zА-Яа-я]+", article)
    return match.group() if match else None


def _parse_sheet(worksheet, sheet_name: str, category_default: str | None, status_default: str) -> list[ParsedItem]:
    headers: list[str | None] = []
    items: list[ParsedItem] = []

    for row_index, row in enumerate(worksheet.iter_rows(min_row=1, max_row=worksheet.max_row), start=1):
        if row_index == 1:
            headers = [_clean(cell.value) for cell in row]
            continue

        values = [cell.value for cell in row]
        article = _clean(values[0]) if values else None
        if not article:
            continue

        fields: dict[str, str] = {}
        specs: dict[str, str] = {}
        for header, raw_value in zip(headers, values):
            if header is None or _PRICE_HEADER_RE.match(header):
                continue
            value = _clean(raw_value)
            if value is None:
                continue
            mapped = _COMMON_FIELD_MAP.get(header)
            if mapped:
                fields[mapped] = value
            else:
                specs[header] = value

        type_field = fields.get("type_field", "")
        is_accessory = bool(_ACCESSORY_TYPE_RE.search(type_field))
        raw_series = fields.get("series")
        series = None if is_accessory else raw_series
        compatible_series = (
            [part.strip() for part in raw_series.split("/") if part.strip()]
            if is_accessory and raw_series
            else []
        )

        status = status_default
        if status_default == "active":
            article_cell = row[0]
            fill = article_cell.fill.fgColor.rgb if article_cell.fill and article_cell.fill.fgColor else None
            if fill == _DISCONTINUED_IN_STOCK_FILL:
                status = "discontinued_in_stock"

        items.append(ParsedItem(
            article=article,
            sheet=sheet_name,
            category_code=_category_code_for_article(article, category_default),
            type_field=type_field,
            name=fields.get("name"),
            series=series,
            size=fields.get("size"),
            is_accessory=is_accessory,
            status=status,
            specs=specs,
            compatible_series=compatible_series,
        ))

    return items


def parse_workbook(path: Path) -> list[ParsedItem]:
    """Parse every known category sheet in the price-list workbook.

    Deliberately NOT read_only=True: discontinued-in-stock detection needs
    cell.fill, which read-only mode doesn't reliably expose. This runs once
    per manual /upload_pricelist, not per user question, so the extra load
    time (~25s for the full ~18k-row workbook, measured) is acceptable —
    call it via asyncio.to_thread() from the bot handler, never inline.
    """
    import openpyxl

    workbook = openpyxl.load_workbook(path, data_only=True)
    try:
        items: list[ParsedItem] = []
        for sheet_name, (category_code, status_default) in CATEGORY_SHEETS.items():
            if sheet_name not in workbook.sheetnames:
                continue
            items.extend(_parse_sheet(workbook[sheet_name], sheet_name, category_code, status_default))
        return items
    finally:
        workbook.close()
